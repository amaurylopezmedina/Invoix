import os
import sys
import json
import time
import calendar
from datetime import datetime, timedelta

import re
import smtplib
import time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import List, Optional

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from cryptography.fernet import Fernet
from openpyxl import Workbook

from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# RUTAS BÁSICAS
# ------------------------------------------------------------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
CONFIG_DIR = os.path.join(BASE_DIR, "config")
os.makedirs(CONFIG_DIR, exist_ok=True)

CONFIG_FILE = os.path.join(CONFIG_DIR, "config_rcorreo.json")
KEY_FILE = os.path.join(CONFIG_DIR, "email_key.key")

# IMPORTS SEGÚN TU ESTRUCTURA
from db.uDB import ConectarDB


class EmailConfig:
    def __init__(
        self,
        smtp_server: str,
        smtp_port: int,
        username: str,
        password: str,
        max_retries: int = 3,
        retry_delay: int = 5,
    ):
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.username = username
        self.password = password
        self.max_retries = max_retries
        self.retry_delay = retry_delay


class EmailSender:
    def __init__(self, config: EmailConfig):
        self.config = config

    @staticmethod
    def validate_email(email: str) -> bool:
        """Valida que el formato del correo electrónico sea correcto."""
        pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
        return bool(re.match(pattern, email))

    def create_message(
        self,
        from_address: str,
        to_addresses: List[str],
        subject: str,
        body: str,
        attachments: Optional[List[str]] = None,
    ) -> MIMEMultipart:
        """Crea el mensaje con sus adjuntos."""
        # Validar correos
        if not self.validate_email(from_address):
            raise ValueError(f"Correo de origen inválido: {from_address}")

        for email in to_addresses:
            if not self.validate_email(email):
                raise ValueError(f"Correo de destino inválido: {email}")

        # Crear mensaje
        msg = MIMEMultipart()
        msg["From"] = from_address
        msg["To"] = ", ".join(to_addresses)
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "html"))

        # Procesar adjuntos
        if attachments:
            for file_path in attachments:
                path = Path(file_path)
                if not path.exists():
                    raise FileNotFoundError(f"No se encontró el archivo: {file_path}")

                try:
                    with open(path, "rb") as attachment:
                        part = MIMEBase("application", "octet-stream")
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            "Content-Disposition", f"attachment; filename= {path.name}"
                        )
                        msg.attach(part)
                except Exception as e:
                    raise Exception(
                        f"Error al procesar el archivo {path.name}: {str(e)}"
                    )

        return msg

    def send_email(
        self,
        from_address: str,
        to_addresses: List[str],
        subject: str,
        body: str,
        attachments: Optional[List[str]] = None,
    ) -> bool:
        """Envía el correo con reintentos en caso de fallo."""
        msg = self.create_message(
            from_address, to_addresses, subject, body, attachments
        )

        for attempt in range(self.config.max_retries):
            try:
                with smtplib.SMTP_SSL(
                    self.config.smtp_server, self.config.smtp_port
                ) as server:
                    server.login(self.config.username, self.config.password)
                    server.send_message(msg)
                    print("Correo enviado exitosamente")
                    return True

            except smtplib.SMTPAuthenticationError:
                raise Exception("Error de autenticación: revise usuario y contraseña")

            except smtplib.SMTPConnectError:
                print(f"Error de conexión en intento {attempt + 1}")
                if attempt < self.config.max_retries - 1:
                    time.sleep(self.config.retry_delay)

            except Exception as e:
                print(f"Error en intento {attempt + 1}: {str(e)}")
                if attempt < self.config.max_retries - 1:
                    time.sleep(self.config.retry_delay)

        raise Exception(
            f"No se pudo enviar el correo después de {self.config.max_retries} intentos"
        )


# ------------------------------------------------------------
# CIFRADO
# ------------------------------------------------------------
def load_key():
    if not os.path.exists(KEY_FILE):
        raise FileNotFoundError(
            f"No se encontró la llave de cifrado: {KEY_FILE}. "
            "Primero configure con ConfigResumenECF.py."
        )
    with open(KEY_FILE, "rb") as f:
        return f.read()


def decrypt_value(token: str, key: bytes) -> str:
    f = Fernet(key)
    return f.decrypt(token.encode("utf-8")).decode("utf-8")


# ------------------------------------------------------------
# CARGA DE CONFIGURACIÓN
# ------------------------------------------------------------
def cargar_config():
    if not os.path.exists(CONFIG_FILE):
        raise FileNotFoundError(
            f"No se encontró el archivo de configuración: {CONFIG_FILE}. "
            "Primero configure con ConfigResumenECF.py."
        )
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


# ------------------------------------------------------------
# HORA Y PERIODICIDAD
# ------------------------------------------------------------
def interpretar_hora(hora_str: str, formato24: bool):
    hora_str = hora_str.strip()

    if formato24:
        # Formato HH:MM
        return datetime.strptime(hora_str, "%H:%M").time()
    else:
        # Formato 12h: 1:30 PM, 01:30 pm, 7pm, etc.
        normalizada = hora_str.upper().replace("AM", " AM").replace("PM", " PM")
        formatos = ["%I:%M %p", "%I %p"]
        for fmt in formatos:
            try:
                return datetime.strptime(normalizada, fmt).time()
            except ValueError:
                continue
        raise ValueError(f"Formato 12h inválido: '{hora_str}'")


def calcular_proxima_ejecucion(hora_str, formato24, periodicidad, ultima_ejecucion):
    ahora = datetime.now()
    hora_obj = interpretar_hora(hora_str, formato24)

    if ultima_ejecucion is None:
        ultima_ejecucion = ahora

    base_date = ultima_ejecucion.date()
    base_day = ultima_ejecucion.day
    periodicidad = (periodicidad or "diario").lower()

    # ---- DIARIO ----
    if periodicidad == "diario":
        proxima = datetime.combine(ahora.date(), hora_obj)
        if proxima <= ahora:
            proxima += timedelta(days=1)
        return proxima

    # ---- SEMANAL ----
    if periodicidad == "semanal":
        proxima = datetime.combine(base_date, hora_obj)
        while proxima <= ahora:
            proxima += timedelta(weeks=1)
        return proxima

    # ---- MENSUAL ----
    if periodicidad == "mensual":
        year = ahora.year
        month = ahora.month
        while True:
            try:
                proxima = datetime(
                    year, month, base_day, hora_obj.hour, hora_obj.minute
                )
            except ValueError:
                # Si no existe ese día en el mes → último del mes
                last_day = calendar.monthrange(year, month)[1]
                proxima = datetime(
                    year, month, last_day, hora_obj.hour, hora_obj.minute
                )

            if proxima > ahora:
                return proxima

            month += 1
            if month > 12:
                month = 1
                year += 1

    # fallback: diario
    return calcular_proxima_ejecucion(hora_str, formato24, "diario", ultima_ejecucion)


# ------------------------------------------------------------
# CONSULTAS A LA BD
# ------------------------------------------------------------
def obtener_resumen(cn, query_resumen: str):
    """
    Ejecuta el query de resumen que debe devolver:
    NumeroFacturaInterna, eNCF, RNCComprador,
    RazonSocialComprador, EstadoFiscal, ResultadoEstadoFiscal
    """
    return cn.fetch_query(query_resumen)


def obtener_totales_estado(cn, query_totales: str):
    """
    Ejecuta el query de totales que debe devolver:
    EstadoFiscal, ResultadoEstadoFiscal, Cantidad
    """
    return cn.fetch_query(query_totales)


def obtener_nombre_archivo(cn, query_nombre_archivo: str) -> str:
    """
    Ejecuta el query que devuelve el nombre del archivo.
    Debe devolver una sola columna (ej: NombreArchivo).
    """
    rows = cn.fetch_query(query_nombre_archivo)
    if not rows:
        fecha = datetime.now().strftime("%Y%m%d")
        nombre = f"Resumen_{fecha}.xlsx"
    else:
        valor = rows[0][0]
        nombre = str(valor) if valor else ""
        if not nombre:
            fecha = datetime.now().strftime("%Y%m%d")
            nombre = f"Resumen_{fecha}.xlsx"

    if not nombre.lower().endswith(".xlsx"):
        nombre += ".xlsx"

    return os.path.join(BASE_DIR, nombre)


def obtener_datos_empresa(cn, query_empresa: str):
    """
    Ejecuta el query que devuelve RNC y NombreEmpresa.
    """
    rows = cn.fetch_query(query_empresa)
    if not rows:
        return "000000000", "Empresa no definida"

    rnc = str(rows[0][0]) if rows[0][0] is not None else "000000000"
    nombre = (
        str(rows[0][1])
        if len(rows[0]) > 1 and rows[0][1] is not None
        else "Empresa sin nombre"
    )
    return rnc, nombre


# ------------------------------------------------------------
# GENERACIÓN DE EXCEL
# ------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.utils import get_column_letter


def generar_excel_resumen(rows_resumen, rows_totales, ruta_archivo: str):
    """
    Genera un Excel completo estilo Dashboard:
    ✔ Hoja Resumen (detalle)
    ✔ Hoja Totales (tablas + gráficos avanzados)
    ✔ Hoja Dashboard (KPI + gráficos interactivos)
    """

    wb = Workbook()

    # ======================================================================
    # ========================== HOJA 1 — DETALLE ==========================
    # ======================================================================
    ws1 = wb.active
    ws1.title = "Resumen eCF"

    headers1 = [
        "NumeroFacturaInterna",
        "eNCF",
        "RNCComprador",
        "RazonSocialComprador",
        "EstadoFiscal",
        "ResultadoEstadoFiscal",
    ]
    ws1.append(headers1)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4A86E8")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Encabezados
    for col in range(1, len(headers1) + 1):
        c = ws1.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = align_center

    # Datos
    for fila in rows_resumen:
        ws1.append(list(fila))

    # Bordes
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, max_col=len(headers1)):
        for cell in row:
            cell.border = thin_border

    # Tabla
    table_ref = f"A1:{chr(64 + len(headers1))}{ws1.max_row}"
    table = Table(displayName="TablaDetalle", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws1.add_table(table)

    # Autoajuste de columnas
    for col in ws1.columns:
        length = max(len(str(cell.value)) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = length + 2

    ws1.freeze_panes = "A2"

    # ======================================================================
    # ========================== HOJA 2 — TOTALES ==========================
    # ======================================================================
    ws2 = wb.create_sheet("Totales")

    headers2 = ["EstadoFiscal", "ResultadoEstadoFiscal", "Cantidad"]
    ws2.append(headers2)

    for col in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = align_center

    total_general = 0
    for estado, resultado, cantidad in rows_totales:
        ws2.append([estado, resultado, cantidad])
        total_general += cantidad

    ws2.append(["", "TOTAL GENERAL", total_general])

    # Bordes
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, max_col=3):
        for cell in row:
            cell.border = thin_border

    # Tabla Totales
    table_ref2 = f"A1:C{ws2.max_row}"
    tabla2 = Table(displayName="TablaTotales", ref=table_ref2)
    tabla2.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True
    )
    ws2.add_table(tabla2)

    # Autoajuste
    for col in ws2.columns:
        length = max(len(str(cell.value)) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = length + 2

    ws2.freeze_panes = "A2"

    # ======================================================================
    # =================== GRÁFICOS AVANZADOS ===============================
    # ======================================================================

    fila_base = ws2.max_row + 3

    # 1) GRÁFICO DE BARRAS
    bar = BarChart()
    bar.title = "Totales por Estado Fiscal"
    datos = Reference(ws2, min_col=3, min_row=1, max_row=ws2.max_row - 1)
    categorias = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row - 1)
    bar.add_data(datos, titles_from_data=True)
    bar.set_categories(categorias)
    bar.height = 10
    bar.width = 20
    ws2.add_chart(bar, f"A{fila_base}")
    fila_base += 15

    # 2) PIE CHART
    pie = PieChart()
    pie.title = "Distribución (%)"
    pie.add_data(datos, titles_from_data=True)
    pie.set_categories(categorias)
    pie.height = 10
    pie.width = 12
    ws2.add_chart(pie, f"A{fila_base}")
    fila_base += 15

    # 3) LINE CHART
    line = LineChart()
    line.title = "Tendencia Visual"
    line.add_data(datos, titles_from_data=True)
    line.set_categories(categorias)
    line.height = 10
    line.width = 20
    ws2.add_chart(line, f"A{fila_base}")
    fila_base += 15

    # 4) BARRAS APILADAS
    bar_stack = BarChart()
    bar_stack.type = "col"
    bar_stack.grouping = "stacked"
    bar_stack.title = "Barras Apiladas"
    bar_stack.add_data(datos, titles_from_data=True)
    bar_stack.set_categories(
        Reference(ws2, min_col=2, min_row=2, max_row=ws2.max_row - 1)
    )
    bar_stack.height = 10
    bar_stack.width = 20
    ws2.add_chart(bar_stack, f"A{fila_base}")
    fila_base += 15

    # 5) GRÁFICO COMBINADO
    combo_bar = BarChart()
    combo_bar.add_data(datos, titles_from_data=True)
    combo_bar.set_categories(categorias)
    combo_bar.title = "Comparación Visual"
    combo_bar.height = 10
    combo_bar.width = 20

    combo_line = LineChart()
    combo_line.add_data(datos, titles_from_data=True)
    combo_line.y_axis.axId = 200

    combo_bar += combo_line
    ws2.add_chart(combo_bar, f"A{fila_base}")
    fila_base += 15

    # ======================================================================
    # ========================== HOJA DASHBOARD =============================
    # ======================================================================
    ws_dash = wb.create_sheet("Dashboard")

    ws_dash["A1"] = "Indicadores Clave (KPI)"
    ws_dash["A1"].font = Font(bold=True, size=16, color="1F4E79")

    # KPIS
    aceptados = sum(c for (e, r, c) in rows_totales if str(e) == "0")
    rechazados = sum(c for (e, r, c) in rows_totales if str(e) == "99")
    porc_exito = (aceptados / total_general * 100) if total_general > 0 else 0

    ws_dash.append(["", ""])
    ws_dash.append(["Total Documentos", total_general])
    ws_dash.append(["Aceptados", aceptados])
    ws_dash.append(["Rechazados", rechazados])
    ws_dash.append(["Tasa de Éxito (%)", round(porc_exito, 2)])

    # Tabla dashboard (para gráficos interactivos)
    ws_dash.append([""])
    ws_dash.append(["EstadoFiscal", "ResultadoEstadoFiscal", "Cantidad"])
    for fila in rows_totales:
        ws_dash.append(list(fila))

    wb.save(ruta_archivo)
    return ruta_archivo


# ------------------------------------------------------------
# FORMATEO DEL CUERPO DEL CORREO (SOLO TOTALES)
# ------------------------------------------------------------
def formatear_totales_html(rows_totales):
    """
    Genera un cuerpo HTML con formato de tabla profesional:
    ✔ Encabezados azules
    ✔ Bordes
    ✔ Totales resaltados
    """

    fecha = (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")

    if not rows_totales:
        return f"""
        <h3>Resumen de Totales e-CF — {fecha}</h3>
        <p>No se encontraron documentos para el día anterior.</p>
        """

    # Encabezado HTML
    html = f"""
    <h2 style="font-family: Arial; color:#333;">
        Resumen de Totales e-CF – {fecha}
    </h2>

    <table style="border-collapse: collapse; width: 100%; font-family: Arial; font-size: 14px;">
        <thead>
            <tr style="background-color:#4A86E8; color:white; text-align:left;">
                <th style="border:1px solid #ccc; padding:8px;">EstadoFiscal</th>
                <th style="border:1px solid #ccc; padding:8px;">ResultadoEstadoFiscal</th>
                <th style="border:1px solid #ccc; padding:8px;">Cantidad</th>
            </tr>
        </thead>
        <tbody>
    """

    total_general = 0

    for estado, resultado, cantidad in rows_totales:
        total_general += cantidad
        html += f"""
            <tr>
                <td style="border:1px solid #ccc; padding:6px;">{estado}</td>
                <td style="border:1px solid #ccc; padding:6px;">{resultado}</td>
                <td style="border:1px solid #ccc; padding:6px; text-align:right;">{cantidad}</td>
            </tr>
        """

    # Total general
    html += f"""
        </tbody>
        <tfoot>
            <tr style="background:#eef;">
                <td colspan="2" style="border:1px solid #ccc; padding:8px;"><b>Total General</b></td>
                <td style="border:1px solid #ccc; padding:8px; text-align:right;"><b>{total_general}</b></td>
            </tr>
        </tfoot>
    </table>
    """

    return html


# ------------------------------------------------------------
# ENVÍO DEL CORREO
# ------------------------------------------------------------
def enviar_correo_resumen(config, key):
    """
    Ejecuta los queries, genera el Excel, arma el cuerpo
    y envía el correo con adjunto.
    """
    cn = ConectarDB()

    query_resumen = config.get("query_resumen", "")
    query_totales = config.get("query_totales_estado", "")
    query_nombre_archivo = config.get("query_nombre_archivo", "")
    query_empresa = config.get("query_empresa", "")

    # Datos
    rows_resumen = obtener_resumen(cn, query_resumen)
    rows_totales = obtener_totales_estado(cn, query_totales)
    ruta_archivo = obtener_nombre_archivo(cn, query_nombre_archivo)
    rnc, empresa = obtener_datos_empresa(cn, query_empresa)

    # Excel con detalle
    generar_excel_resumen(rows_resumen, rows_totales, ruta_archivo)

    # Cuerpo = totales
    body = formatear_totales_html(rows_totales)

    # Config correo
    c = config["correo"]
    if c.get("password_encrypted"):
        password_real = decrypt_value(c["password_encrypted"], key)
    else:
        password_real = c.get("password", "")

    email_cfg = EmailConfig(
        smtp_server=c["smtp_server"],
        smtp_port=c["smtp_port"],
        username=c["username"],
        password=password_real,
    )

    sender = EmailSender(email_cfg)

    fecha = (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")
    subject = f"Resumen e-CF — {empresa} (RNC {rnc}) — {fecha}"

    sender.send_email(
        from_address=c["from_address"],
        to_addresses=c["to_addresses"],
        subject=subject,
        body=body,
        attachments=[ruta_archivo],
    )

    print(f"[OK] Correo enviado con adjunto: {ruta_archivo}")


# ------------------------------------------------------------
# SERVICIO PRINCIPAL
# ------------------------------------------------------------
def servicio():
    print("[INFO] Servicio Resumen eCF iniciado...")
    key = load_key()
    config = cargar_config()

    ultima_ejecucion = None

    # Enviar al arrancar si aplicar
    if config.get("enviar_al_entrar", False):
        try:
            print("[INFO] Enviando correo inicial (enviar_al_entrar = true)...")
            enviar_correo_resumen(config, key)
            ultima_ejecucion = datetime.now()
        except Exception as e:
            print(f"[ERROR] Error al enviar correo inicial: {e}")

    # Bucle infinito del servicio
    while True:
        # Siempre recargar config por si se cambia mientras corre
        config = cargar_config()

        hora_envio = config.get("hora_envio", "01:00")
        formato24 = config.get("formato24", True)
        periodicidad = config.get("periodicidad", "diario")

        try:
            proxima = calcular_proxima_ejecucion(
                hora_envio, formato24, periodicidad, ultima_ejecucion
            )
        except Exception as e:
            print(f"[ERROR] Hora/periodicidad inválida en configuración: {e}")
            # Esperar 60s y reintentar
            time.sleep(60)
            continue

        print(f"[INFO] Próxima ejecución programada para: {proxima}")

        espera = (proxima - datetime.now()).total_seconds()
        if espera > 0:
            time.sleep(espera)

        print("[INFO] Ejecutando envío programado...")

        try:
            enviar_correo_resumen(config, key)
            ultima_ejecucion = datetime.now()
        except Exception as e:
            print(f"[ERROR] Error al enviar correo programado: {e}")
            # Evitar loop loco: pequeña espera
            time.sleep(60)


if __name__ == "__main__":
    servicio()
