import os
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# ===============================
# VARIABLES GLOBALES
# ===============================
comprobantes = []
directorio_xml = ""
resultados = []
resumen = {"total": 0, "ok": 0, "no": 0}


# ===============================
# CARGAR EXCEL
# ===============================
def cargar_excel():
    global comprobantes
    path = filedialog.askopenfilename(
        title="Seleccionar Excel de comprobantes", filetypes=[("Excel files", "*.xlsx")]
    )
    if not path:
        return

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    comprobantes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            comprobantes.append(str(row[0]).strip())

    messagebox.showinfo(
        "Excel cargado", f"✔ Comprobantes cargados: {len(comprobantes)}"
    )


# ===============================
# DIRECTORIO XML
# ===============================
def seleccionar_directorio():
    global directorio_xml
    directorio_xml = filedialog.askdirectory(title="Seleccionar carpeta de XML")
    if directorio_xml:
        messagebox.showinfo("Directorio seleccionado", directorio_xml)


# ===============================
# XML
# ===============================
def obtener_signature_6(xml_path):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        for elem in root.iter():
            if elem.tag.endswith("SignatureValue"):
                return elem.text.strip()[:6]
        return ""
    except:
        return "ERROR_XML"


def extraer_rnc(nombre, encf):
    return nombre.replace(encf, "").replace(".xml", "")


# ===============================
# EJECUTAR COMPARACIÓN
# ===============================
def ejecutar_comparacion():
    global resultados, resumen

    if not comprobantes:
        messagebox.showerror("Error", "Debe cargar el Excel de comprobantes")
        return
    if not directorio_xml:
        messagebox.showerror("Error", "Debe seleccionar el directorio de XML")
        return

    resultados = []
    resumen = {"total": len(comprobantes), "ok": 0, "no": 0}

    resultado_txt.delete(1.0, tk.END)
    resultado_txt.insert(tk.END, "RNC\tComprobante\tSignature6\tEstado\n")
    resultado_txt.insert(tk.END, "-" * 90 + "\n")

    archivos = os.listdir(directorio_xml)

    for encf in comprobantes:
        encontrado = False

        for file in archivos:
            if encf in file and file.lower().endswith(".xml"):
                sig = obtener_signature_6(os.path.join(directorio_xml, file))
                rnc = extraer_rnc(file, encf)
                resultados.append([rnc, encf, sig, file, "OK"])
                resultado_txt.insert(tk.END, f"{rnc}\t{encf}\t{sig}\tOK\n")
                resumen["ok"] += 1
                encontrado = True
                break

        if not encontrado:
            resultados.append(["", encf, "", "", "NO_ENCONTRADO"])
            resultado_txt.insert(tk.END, f"\t{encf}\t\tNO_ENCONTRADO\n")
            resumen["no"] += 1

    messagebox.showinfo(
        "Comparación finalizada",
        f"✔ Comparación completada\n\n"
        f"Total: {resumen['total']}\n"
        f"OK: {resumen['ok']}\n"
        f"NO_ENCONTRADO: {resumen['no']}",
    )


# ===============================
# EXPORTAR
# ===============================
def exportar():
    if not resultados:
        messagebox.showerror("Error", "Debe ejecutar la comparación antes de exportar")
        return

    excel_out = filedialog.asksaveasfilename(
        title="Guardar resultado",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile="Resultado_Signature_eCF.xlsx",
    )
    if not excel_out:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ws.append(["RNC", "Comprobante", "Signature6", "Archivo", "Estado"])

    verde = PatternFill("solid", fgColor="C6EFCE")
    rojo = PatternFill("solid", fgColor="FFC7CE")

    for r in resultados:
        ws.append(r)
        fill = verde if r[4] == "OK" else rojo
        for c in ws[ws.max_row]:
            c.fill = fill

    ws_sum = wb.create_sheet("Resumen")
    ws_sum.append(["Concepto", "Cantidad"])
    ws_sum.append(["Total procesados", resumen["total"]])
    ws_sum.append(["OK", resumen["ok"]])
    ws_sum.append(["NO_ENCONTRADO", resumen["no"]])

    wb.save(excel_out)

    csv_out = excel_out.replace(".xlsx", ".csv")
    with open(csv_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["RNC", "Comprobante", "Signature6", "Archivo", "Estado"])
        writer.writerows(resultados)

    messagebox.showinfo(
        "Exportación completada",
        f"✔ Archivos generados\n\nExcel:\n{excel_out}\nCSV:\n{csv_out}",
    )


# ===============================
# UI
# ===============================
root = tk.Tk()
root.title("Extractor SignatureValue e-CF – UI Profesional")
root.geometry("1050x580")

frame = tk.Frame(root)
frame.pack(pady=10)

tk.Button(frame, text="📊 Cargar Excel", width=20, command=cargar_excel).grid(
    row=0, column=0, padx=5
)
tk.Button(
    frame, text="📁 Directorio XML", width=24, command=seleccionar_directorio
).grid(row=0, column=1, padx=5)
tk.Button(
    frame, text="▶ Ejecutar Comparación", width=26, command=ejecutar_comparacion
).grid(row=0, column=2, padx=5)
tk.Button(frame, text="💾 Exportar", width=18, command=exportar).grid(
    row=0, column=3, padx=5
)

resultado_txt = scrolledtext.ScrolledText(root, font=("Consolas", 10))
resultado_txt.pack(expand=True, fill="both", padx=10, pady=10)

root.mainloop()
